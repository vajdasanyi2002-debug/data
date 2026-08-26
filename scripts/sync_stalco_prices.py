import os
import sys
import time
from pathlib import Path

import openpyxl
import requests


# ======================================================
# CONFIG
# ======================================================

EXCEL_PATH = Path(
    "Date_produse_Stalco.xlsx"
)

WEBHOOK_URL = os.environ[
    "ODOO_WEBHOOK_URL"
]

BATCH_SIZE = 200
TIMEOUT = 120


# ======================================================
# TEST / REPLICATION SETTINGS
# ======================================================

TEST_BARCODE = "5901466110942"

# False:
# csak Excel -> Hala teszt
#
# True:
# Hala összes Stalco variant ára
# -> Precizia + Lunca
#
REPLICATE_STORE_PRICES = False


# ======================================================
# HELPERS
# ======================================================

def blank(value):
    return value is None or (
        isinstance(value, str)
        and not value.strip()
    )


def barcode_value(value):

    if blank(value):
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):

        if value.is_integer():
            return str(
                int(value)
            )

        return None

    value = str(
        value
    ).strip()

    if (
        value.endswith(".0")
        and value[:-2].isdigit()
    ):
        value = value[:-2]

    return value or None


def number_value(value):

    if blank(value):
        return None

    if isinstance(value, str):

        value = (
            value
            .strip()
            .replace(",", ".")
        )

    return float(value)


def text_value(value):

    if blank(value):
        return None

    if isinstance(value, int):
        return str(value)

    if (
        isinstance(value, float)
        and value.is_integer()
    ):
        return str(
            int(value)
        )

    value = str(
        value
    ).strip()

    if (
        value.endswith(".0")
        and value[:-2].isdigit()
    ):
        value = value[:-2]

    return value or None


# ======================================================
# LOAD EXCEL
# ======================================================

def load_items():

    wb = openpyxl.load_workbook(
        EXCEL_PATH,
        read_only=True,
        data_only=True,
    )

    ws = wb.active

    headers = [
        cell.value
        for cell in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]


    required = [
        "Cod e bare/Cod furnizo",
        "Denumire Stalco engleza",
        "UM",
        "PA Euro",
        "PV RON cu TVA",
        "weight",
        "l10n_ro_net_weight",
        "hs_code",
    ]


    for name in required:

        if name not in headers:

            raise RuntimeError(
                "Missing Excel column: %s"
                % name
            )


    columns = {
        name:
            headers.index(name)
        for name in headers
    }


    items = []
    seen = set()

    skipped_barcode = 0
    skipped_empty = 0


    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        # ----------------------------------------------
        # BARCODE
        # ----------------------------------------------

        barcode = barcode_value(
            row[
                columns[
                    "Cod e bare/Cod furnizo"
                ]
            ]
        )

        if not barcode:

            skipped_barcode += 1

            continue


        if barcode in seen:

            raise RuntimeError(
                "Duplicate barcode in Excel: %s"
                % barcode
            )


        seen.add(barcode)


        item = {
            "barcode":
                barcode,

            "supplier_code":
                barcode,
        }


        # ----------------------------------------------
        # SUPPLIER NAME
        # ----------------------------------------------

        supplier_name = text_value(
            row[
                columns[
                    "Denumire Stalco engleza"
                ]
            ]
        )

        if supplier_name is not None:

            item[
                "supplier_name"
            ] = supplier_name


        # ----------------------------------------------
        # PURCHASE PRICE EUR
        # ----------------------------------------------

        supplier_price = number_value(
            row[
                columns[
                    "PA Euro"
                ]
            ]
        )

        if supplier_price is not None:

            item[
                "supplier_price"
            ] = supplier_price


        # ----------------------------------------------
        # HALA SALE PRICE
        # ----------------------------------------------

        sale_price_hala = number_value(
            row[
                columns[
                    "PV RON cu TVA"
                ]
            ]
        )

        if sale_price_hala is not None:

            item[
                "sale_price_hala"
            ] = sale_price_hala


        # ----------------------------------------------
        # PURCHASE UOM
        # ----------------------------------------------

        um = text_value(
            row[
                columns[
                    "UM"
                ]
            ]
        )

        if um is not None:

            item["um"] = um


        # ----------------------------------------------
        # GROSS WEIGHT
        # ----------------------------------------------

        weight = number_value(
            row[
                columns[
                    "weight"
                ]
            ]
        )

        if weight is not None:

            item[
                "weight"
            ] = weight


        # ----------------------------------------------
        # NET WEIGHT
        # ----------------------------------------------

        net_weight = number_value(
            row[
                columns[
                    "l10n_ro_net_weight"
                ]
            ]
        )

        if net_weight is not None:

            item[
                "l10n_ro_net_weight"
            ] = net_weight


        # ----------------------------------------------
        # HS CODE
        # ----------------------------------------------

        hs_code = text_value(
            row[
                columns[
                    "hs_code"
                ]
            ]
        )

        if hs_code is not None:

            item[
                "hs_code"
            ] = hs_code


        if len(item) <= 2:

            skipped_empty += 1

            continue


        items.append(item)


    wb.close()


    print(
        "Prepared:",
        len(items)
    )

    print(
        "Skipped without barcode:",
        skipped_barcode
    )

    print(
        "Skipped without data:",
        skipped_empty
    )


    return items


# ======================================================
# SEND BATCH
# ======================================================

def send_batch(
    batch,
    batch_number,
    total_batches,
    replicate_store_prices=False,
):

    response = requests.post(
        WEBHOOK_URL,
        json={
            "source":
                "github_stalco_product_data",

            "items":
                batch,

            "replicate_store_prices":
                replicate_store_prices,
        },
        timeout=TIMEOUT,
    )


    if not response.ok:

        raise RuntimeError(
            "Batch %s/%s failed: "
            "HTTP %s - %s"
            % (
                batch_number,
                total_batches,
                response.status_code,
                response.text[:1000],
            )
        )


    print(
        "Batch %s/%s OK - %s products"
        % (
            batch_number,
            total_batches,
            len(batch),
        )
    )


# ======================================================
# MAIN
# ======================================================

def main():

    items = load_items()


    # ==================================================
    # TEST MODE
    # ==================================================

    if TEST_BARCODE:

        items = [
            item
            for item in items
            if item.get("barcode")
            == TEST_BARCODE
        ]

        if not items:

            raise RuntimeError(
                "Test product not found in Excel: %s"
                % TEST_BARCODE
            )

        print(
            "TEST MODE - barcode:",
            TEST_BARCODE,
        )

        print(
            "TEST PAYLOAD:",
            items[0],
        )


    total_batches = (
        len(items)
        + BATCH_SIZE
        - 1
    ) // BATCH_SIZE


    for start in range(
        0,
        len(items),
        BATCH_SIZE,
    ):

        batch = items[
            start:
            start + BATCH_SIZE
        ]

        batch_number = (
            start // BATCH_SIZE
        ) + 1


        # A teljes 402 -> 403/404 replikációt
        # CSAK AZ UTOLSÓ batch után kérjük.
        #
        # Így nem fut le 50-szer egy teljes importnál.

        is_last_batch = (
            batch_number
            == total_batches
        )

        do_replication = (
            REPLICATE_STORE_PRICES
            and is_last_batch
        )


        send_batch(
            batch,
            batch_number,
            total_batches,
            replicate_store_prices=do_replication,
        )


        time.sleep(0.15)


    print(
        "DONE - %s rows sent to Odoo."
        % len(items)
    )

    print(
        "Store price replication:",
        REPLICATE_STORE_PRICES
    )


# ======================================================
# ENTRY
# ======================================================

if __name__ == "__main__":

    try:

        main()

    except Exception as exc:

        print(
            "ERROR:",
            exc,
            file=sys.stderr,
        )

        sys.exit(1)
